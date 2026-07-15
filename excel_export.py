"""
excel_export.py — CalcVault (Ramboll Edition)
=============================================
Owner-side utility: dump the activity log to a branded .xlsx.

Public API:
    export_activity_log(rows: list[dict],
                        title: str = "CalcVault Activity Log") -> bytes
    export_history(rows: list[dict]) -> bytes         # user history dump
    export_pump_databank(rows: list[dict]) -> bytes   # pump reference dump

Design:
    • Pure in-memory (returns bytes for send_file).
    • One shared _write_sheet() helper — no duplicated styling code.
    • Ramboll blue header row, frozen top row + freeze first column,
      auto-fit column widths, subtle zebra striping, ISO dates.
    • No pandas dependency — openpyxl only, keeping the export path fast.
"""

from __future__ import annotations
import io
import json
from datetime import datetime, date
from typing import List, Dict, Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Font, PatternFill, Border, Side, NamedStyle,
)
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------
# Ramboll palette (Excel expects ARGB hex, no leading '#')
# ---------------------------------------------------------------
BLUE       = "FF1F3F89"
CYAN       = "FF00A0DC"
ROW_ALT    = "FFF3F6FB"
BORDER_CLR = "FFD9E1EF"
INK_DARK   = "FF111827"
WHITE      = "FFFFFFFF"

_thin       = Side(style="thin", color=BORDER_CLR)
_cell_bdr   = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

_HDR_FONT   = Font(name="Calibri", size=11, bold=True, color=WHITE)
_HDR_FILL   = PatternFill("solid", fgColor=BLUE)
_HDR_ALIGN  = Alignment(horizontal="left", vertical="center", wrap_text=True)

_BODY_FONT  = Font(name="Calibri", size=10, color=INK_DARK)
_ALT_FILL   = PatternFill("solid", fgColor=ROW_ALT)
_BODY_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ===============================================================
# Value normaliser  (Excel can't natively serialize dicts/lists)
# ===============================================================
def _norm(v: Any) -> Any:
    if v is None:
        return ""
    if isinstance(v, (dict, list, tuple)):
        # Compact JSON – keeps details readable in the cell
        return json.dumps(v, ensure_ascii=False, default=str)
    if isinstance(v, (datetime, date)):
        return v.isoformat(sep=" ", timespec="seconds") \
               if isinstance(v, datetime) else v.isoformat()
    if isinstance(v, bool):
        return "Yes" if v else "No"
    return v


# ===============================================================
# Sheet writer  (shared by every exporter)
# ===============================================================
def _write_sheet(wb: Workbook,
                 sheet_name: str,
                 rows: Iterable[Dict[str, Any]],
                 preferred_order: List[str] | None = None) -> None:

    rows = list(rows)
    ws   = wb.active if len(wb.worksheets) == 1 and \
                       wb.active.title == "Sheet" else wb.create_sheet()
    ws.title = sheet_name[:31] or "Sheet"

    # ---- Determine columns -------------------------------------------
    if rows:
        seen: List[str] = []
        # Preferred columns first (if present in data)
        for k in preferred_order or []:
            if any(k in r for r in rows) and k not in seen:
                seen.append(k)
        # Then any remaining keys, in order of first appearance
        for r in rows:
            for k in r.keys():
                if k not in seen:
                    seen.append(k)
        headers = seen
    else:
        headers = ["(no data)"]

    # ---- Header row --------------------------------------------------
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c,
                       value=h.replace("_", " ").title())
        cell.font      = _HDR_FONT
        cell.fill      = _HDR_FILL
        cell.alignment = _HDR_ALIGN
        cell.border    = _cell_bdr

    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "B2"   # freeze first row AND first column

    # ---- Body rows ---------------------------------------------------
    max_len = {h: len(h) for h in headers}
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, key in enumerate(headers, start=1):
            v = _norm(row.get(key))
            cell = ws.cell(row=r_idx, column=c_idx, value=v)
            cell.font      = _BODY_FONT
            cell.alignment = _BODY_ALIGN
            cell.border    = _cell_bdr
            if r_idx % 2 == 0:
                cell.fill = _ALT_FILL
            # Track for column width
            length = len(str(v)) if v is not None else 0
            if length > max_len[key]:
                max_len[key] = length

    # ---- Auto-fit column widths  (cap at 60 chars) -------------------
    for c_idx, key in enumerate(headers, start=1):
        width = min(max(max_len[key] + 4, 12), 60)
        ws.column_dimensions[get_column_letter(c_idx)].width = width

    # ---- Autofilter over the whole table -----------------------------
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}" \
                         f"{max(len(rows) + 1, 2)}"


# ===============================================================
# Public exporters
# ===============================================================
_LOG_ORDER = [
    "id", "created_at", "user", "role", "module", "action",
    "status", "report_id", "inputs", "results", "notes",
]

def export_activity_log(rows: List[Dict[str, Any]],
                        title: str = "CalcVault Activity Log") -> bytes:
    """Owner: full audit log dump."""
    wb = Workbook()
    _write_sheet(wb, "Activity Log", rows, preferred_order=_LOG_ORDER)
    _stamp(wb, title)
    return _to_bytes(wb)


_HIST_ORDER = [
    "id", "created_at", "module", "status", "report_id",
    "inputs", "results",
]

def export_history(rows: List[Dict[str, Any]]) -> bytes:
    """User: 'my history' dump."""
    wb = Workbook()
    _write_sheet(wb, "My Calculations", rows, preferred_order=_HIST_ORDER)
    _stamp(wb, "CalcVault — My History")
    return _to_bytes(wb)


_PUMP_ORDER = [
    "id", "vendor", "model", "flow_m3h", "head_m",
    "pump_eff_pct", "motor_eff_pct", "motor_kw",
    "pump_weight_kg", "motor_weight_kg", "notes", "pdf_filename",
]

def export_pump_databank(rows: List[Dict[str, Any]]) -> bytes:
    """Owner: reference pump list dump."""
    wb = Workbook()
    _write_sheet(wb, "Pump Databank", rows, preferred_order=_PUMP_ORDER)
    _stamp(wb, "CalcVault — Pump Databank")
    return _to_bytes(wb)


# ===============================================================
# Metadata + serialisation
# ===============================================================
def _stamp(wb: Workbook, title: str) -> None:
    wb.properties.title    = title
    wb.properties.creator  = "Ramboll CalcVault"
    wb.properties.subject  = "Engineering audit export"
    wb.properties.created  = datetime.now()


def _to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ===============================================================
# Smoke test
# ===============================================================
if __name__ == "__main__":
    sample = [
        {"id": 1, "created_at": datetime.now(), "user": "alice",
         "role": "user", "module": "Pipe Head Loss",
         "action": "run", "status": "approved",
         "report_id": "RPT-20260714-101010-ABCD",
         "inputs":  {"Q": 212, "C": 140, "ID_mm": 210.1, "L_m": 131},
         "results": {"head_loss_m": 1.5657}},
        {"id": 2, "created_at": datetime.now(), "user": "bob",
         "role": "user", "module": "Pipe Diameter",
         "action": "run", "status": "pending",
         "inputs":  {"Q": 40, "V": 1.62},
         "results": {"required_id_mm": 93.4}},
    ]
    with open("_test_activity_log.xlsx", "wb") as f:
        f.write(export_activity_log(sample))
    print("✅ activity log written")